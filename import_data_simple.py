"""
Simple data import script using openpyxl (no pandas required).
"""
import os
import sys
import openpyxl
from decimal import Decimal
from datetime import datetime

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'credit_system.settings')

import django
django.setup()

from loans.models import Customer, Loan
from django.utils.dateparse import parse_date


def parse_date_flexible(date_val):
    """Parse date from various formats."""
    if not date_val:
        return None
    
    if isinstance(date_val, datetime):
        return date_val.date()
    
    if isinstance(date_val, str):
        # Try different formats
        for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                return datetime.strptime(date_val, fmt).date()
            except:
                continue
    
    # Try Django's parser as fallback
    return parse_date(str(date_val))


def import_customer_data():
    """Import customer data from Excel file."""
    file_path = 'customer_data.xlsx'
    
    print(f"Importing customer data from {file_path}...")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        customers_created = 0
        customers_updated = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # Create dict from row
                data = dict(zip(headers, row))
                
                customer_id = data.get('Customer ID')
                if not customer_id:
                    continue
                
                first_name = data.get('First Name', '')
                last_name = data.get('Last Name', '')
                age = data.get('Age', 25)
                phone_number = data.get('Phone Number', '')
                monthly_salary = data.get('Monthly Salary', 0)
                approved_limit = data.get('Approved Limit', 0)
                
                defaults = {
                    'first_name': str(first_name) if first_name else '',
                    'last_name': str(last_name) if last_name else '',
                    'age': int(age) if age else 25,
                    'phone_number': str(int(phone_number)) if phone_number else '',
                    'monthly_salary': Decimal(str(monthly_salary)) if monthly_salary else Decimal('0'),
                    'approved_limit': Decimal(str(approved_limit)) if approved_limit else Decimal('0'),
                }
                
                customer, created = Customer.objects.update_or_create(
                    customer_id=int(customer_id),
                    defaults=defaults
                )
                
                if created:
                    customers_created += 1
                else:
                    customers_updated += 1
                    
            except Exception as e:
                print(f"Error processing customer row: {e}")
                continue
        
        print(f"Customer data imported: {customers_created} created, {customers_updated} updated")
        return True
        
    except Exception as e:
        print(f"Error importing customer data: {e}")
        return False


def import_loan_data():
    """Import loan data from Excel file."""
    file_path = 'loan_data.xlsx'
    
    print(f"Importing loan data from {file_path}...")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        loans_created = 0
        loans_updated = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # Create dict from row
                data = dict(zip(headers, row))
                
                customer_id = data.get('Customer ID')
                loan_id = data.get('Loan ID')
                
                if not customer_id or not loan_id:
                    continue
                
                try:
                    customer = Customer.objects.get(customer_id=int(customer_id))
                except Customer.DoesNotExist:
                    print(f"Customer {customer_id} not found, skipping loan {loan_id}")
                    continue
                
                start_date = parse_date_flexible(data.get('Date of Approval'))
                end_date = parse_date_flexible(data.get('End Date'))
                
                defaults = {
                    'loan_amount': Decimal(str(data.get('Loan Amount', 0))) if data.get('Loan Amount') else Decimal('0'),
                    'tenure': int(data.get('Tenure', 0)) if data.get('Tenure') else 0,
                    'interest_rate': Decimal(str(data.get('Interest Rate', 0))) if data.get('Interest Rate') else Decimal('0'),
                    'monthly_repayment': Decimal(str(data.get('Monthly payment', 0))) if data.get('Monthly payment') else Decimal('0'),
                    'emis_paid_on_time': int(data.get('EMIs paid on Time', 0)) if data.get('EMIs paid on Time') else 0,
                    'start_date': start_date,
                    'end_date': end_date,
                    'is_approved': True,
                }
                
                loan, created = Loan.objects.update_or_create(
                    loan_id=int(loan_id),
                    customer=customer,
                    defaults=defaults
                )
                
                if created:
                    loans_created += 1
                else:
                    loans_updated += 1
                        
            except Exception as e:
                print(f"Error processing loan row: {e}")
                continue
        
        print(f"Loan data imported: {loans_created} created, {loans_updated} updated")
        return True
        
    except Exception as e:
        print(f"Error importing loan data: {e}")
        return False


if __name__ == '__main__':
    print("Starting data import...")
    
    customer_result = import_customer_data()
    loan_result = import_loan_data()
    
    if customer_result and loan_result:
        print("Data import completed successfully!")
    else:
        print("Data import completed with errors.")
        sys.exit(1)
